
import datetime
import os
import pathlib
import platform
import subprocess
from typing import Any, Iterable, Protocol, TypedDict, Generator, overload, get_type_hints, Callable
import warnings

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy as sp
import uncertainties
from matplotlib import colors, ticker
from matplotlib.figure import Figure
from matplotlib.axes import Axes
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.gridspec import GridSpec
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from numpy import ndarray
from numpy import typing as npt
from pandas.core.frame import DataFrame
from scipy import odr
from scipy.signal import find_peaks, savgol_filter
from uncertainties import ufloat, ufloat_fromstr
from uncertainties.core import Variable
from openpyxl import load_workbook

__version__ = "2024.07.07"

def versions() -> list[tuple[str, str]]:
    """Return list of used packages and their versions.
    """
    return [
        ("NumPy", np.version.version),
        ("SciPy", sp.__version__),
        ("Pandas", pd.__version__),
        ("matplotlib", mpl.__version__),
        ("uncertainties", uncertainties.__version__),
        ("photoacoustics", __version__),
    ]


###############
# Constants
###############

# Repeats is not included as metadata in the file.
ATTR_REPEATS = "__PA_REPEATS__"

ATTR_TIME_UNITS = "__PA_TIME_UNITS__"
ATTR_SIGNAL_UNITS = "__PA_SIGNAL_UNITS__"

# Metadata float attributes
ATTRS_FLOAT  = ("Start", "Stop", "Step", "Wavelength", "Bandwidth")
# Metadata integer attributes
ATTRS_INT = ("Averages", )
# Metadata uncertainty attributes
ATTRS_UNC = ("Laser energy before", "Laser energy after")

UFLOAT0 = ufloat(0, 0)
UFLOAT_NAN = ufloat(np.nan, np.nan)

DEFAULT_OPTIONS = {
    "savgol_window_length": 51,
    "savgol_polyorder": 3,
    "on_progress": print,
    "on_error": print,
    "plot_time_trace_rep": True,
}

#################
# Typing related
#################

Array = npt.NDArray[np.float64]

class ReadLiner(Protocol):

    def readline(self) -> str:
        ...

####
# Unused for now
class OAColumns(TypedDict):
    time: float
    signal: float
    
class OAColumnsSmooth(OAColumns):
    signal_smooth: float

class Metadata(TypedDict):
    start: float
    stop: float
    step: float
    wavelength: float
    bandwidth: float
    averages: int
    laser_before: Variable
    laser_after: Variable

# end unused

class TraceAnalysis(TypedDict):
    energy: Variable

    time_peak1: Variable
    signal_peak1: Variable

    time_peak2: Variable
    signal_peak2: Variable

    time_delta: Variable
    signal_delta: Variable


class FileAnalysis(TypedDict):
    path: str
    description: str
    comment: str
    wavelength: float
    bandwith: float
    averages: int
    repeats: int | None

    energy: Variable

    time_delta: Variable
    signal_delta: Variable


class PowerscanAnalysis(TypedDict):
    folder: str
    sam_ref: str
    exc_wavelength: float
    description: str
    slope: Variable
    intercept: Variable
    slope0: Variable


class ExperimentAnalysis(TypedDict):
    abs_sam: float
    abs_ref: float
    ref: str
    sam: str
    exc_wavelength: float
    alpha: Variable


class Options(TypedDict):
    savgol_window_length: int
    savgol_polyorder: int
    on_progress: Callable[[str,], None]
    on_error: Callable[[str,], None]
    plot_time_trace_rep: bool

###################
# Helper functions
###################

def _argmedian_at_t0(signals: list[tuple[Array, Array]]) -> int:
    """Returns the indices of the median value.
    """
    peak_signal: list[float] = [
        signal[np.searchsorted(time, 0)+1] 
        for time, signal in signals
    ]
    el: float = np.percentile(peak_signal, 50, method='closest_observation') # type: ignore
    return peak_signal.index(el)


def get_unc_keys(klass: type) -> tuple[str, ...]:
    """Returns the attribute names annotated as Variable (uncertainty).
    """
    return tuple(k for k, v in get_type_hints(klass).items() if v is Variable)


def unzip_unc_column(df: DataFrame, *column_names: str, drop_unc: bool=False) -> pd.DataFrame:
    """Unzip uncertainty column (`name`) into nominal_value (`name`) and std_dev ()`name_unc`).

    If `drop_unc` is True, only the nominal value will be extracted.
    """

    original_columns = df.columns
    for column_name in column_names:
        if drop_unc:
            df[[column_name, ]] = df[column_name].apply(
                lambda x: (x.nominal_value, )
            ).to_list()
        else:
            df[[column_name, column_name + "_unc"]] = df[column_name].apply(
                lambda x: (x.nominal_value, x.std_dev)
            ).to_list()

    if drop_unc:
        return df
    
    new_columns = []
    for column_name in original_columns:
        new_columns.append(column_name)
        if column_name in column_names:
            new_columns.append(column_name + "_unc")

    return df.reindex(columns=new_columns)


def to_unc_str(nominal_value: float, std_dev: float, units: str | None = None) -> str:
    """Build nice string of a quantity with uncertainty.
    """
    if units is None:
        return "${:.2uL}$".format(ufloat(nominal_value, std_dev))
    return "$({:.2uL})~{}$".format(ufloat(nominal_value, std_dev), units)    

def to_unc_tuple(nominal_values: Iterable[float], std_devs: Iterable[float]) -> tuple[Variable, ...]:
    """Zip iterable of nominal values and std devs into a tuple of uncertainties.
    """
    return tuple(ufloat(v, u) for v, u in zip(nominal_values, std_devs))

def split_unc_tuple(*variables: Variable, container: type=tuple) -> tuple[tuple[float, ...], tuple[float, ...]]:
    """Unzip iterable of uncertainties into a tuple of nominal value and a tuple of std dev
    """
    return container(v.nominal_value for v in variables), container(v.std_dev for v in variables)


def ufloat_nanmean(*variables: Variable) -> Variable:
    """Calculate new uncertainty by averaging an iterable of uncertainties.
    """
    els = [el for el in variables 
             if np.isfinite(el.nominal_value) and np.isfinite(el.std_dev)]
    
    N = len(els)
    if N == 0:
        return UFLOAT_NAN
    
    mx = np.asarray([el.nominal_value for el in els])
    sx =  np.asarray([el.std_dev for el in els])

    if np.all(sx == 0):
        return ufloat(mx.mean(), mx.std())

    wx =  1 / sx / sx
    
    return ufloat((mx * wx).sum() / wx.sum(), np.sqrt(1 / wx.sum()))

def ztest(unc1: Variable, unc2: Variable) -> float:
    """Given two uncertainties

    Compare the means of two samples to see if it is feasible that 
    they come from the same population. 
    The null hypothesis is: the population means are equal.
    
    A very small p-value means that such an extreme observed outcome 
    would be very unlikely under the null hypothesis.

    """
    z = (unc1.nominal_value - unc2.nominal_value) ** 2
    z = z / (unc1.std_dev ** 2 + unc2.std_dev ** 2)
    return sp.stats.norm.sf(np.sqrt(abs(z)))

################
# I/O functions
################


def reorganize_sheets(path: pathlib.Path):
    """Reorganize sheets in an excel file
    """
    wb = load_workbook(path)
    sheetnames = wb.sheetnames
    for ndx, sheetname in enumerate(sheetnames, 0):
        wb.move_sheet(sheetname, -ndx)

    sheetnames = wb.sheetnames
    for ndx, sheetname in enumerate(sheetnames, 0):
        if sheetname.startswith("ref") or sheetname.startswith("sam"):
            wb.move_sheet(sheetname, -ndx+2)

    wb.save(path)


def read_metadata(fi: ReadLiner) -> dict[str, str]:
    """Consume metadata from an open file.
    """
    metadata: dict[str, str] = {}
    cnt = 0
    while  True:
        line = str.strip(fi.readline())
        if not line:
            cnt += 1
            if cnt == 2:
                # Two empty lines
                break
            continue
        key, value = line.split(",", 1)
        metadata[key] = value

    return metadata


@overload
def parse_metadata_content(metadata: Any, repeats: None) -> Generator[tuple[str, int | float | Variable | str], None, None]:
    ...

@overload
def parse_metadata_content(metadata: Any, repeats: int) -> Generator[tuple[str, tuple[int, ...] | tuple[float, ...] | tuple[Variable, ...] | tuple[str, ...]], None, None]:
    ...

def parse_metadata_content(metadata: dict[str, str], repeats: int | None=None) -> Generator[tuple[Any, Any], None, None]:
    """Parse metadata content.
    """

    if repeats is None:
        fun = lambda conv, x: conv(x) # type: ignore
    else:
        fun = lambda conv, x: tuple(conv(el) for el in x.split(",")) # type: ignore

    for k, v in metadata.items():
        if k in ATTRS_FLOAT:
            yield k, fun(float, v)
        elif k in ATTRS_INT:
            yield k, fun(int, v)
        elif k in ATTRS_UNC:
            yield k, fun(ufloat_fromstr, v)
        else:
            yield k, fun(str, v)


def read_without_repeats(p: pathlib.Path | str) -> DataFrame:
    """Read Edinburgh Instruments ascii file (with no repeats).
    """
    if isinstance(p, str):
        p = pathlib.Path(p)
        
    with p.open("r", encoding="ascii") as fi:
        metadata = read_metadata(fi)

        # TODO: is the time always in ms? is the signal always in Volts?
        df = pd.read_csv( # type: ignore
            fi, 
            sep=",",
            header=0,
            names=("time", "signal")
        )
        df["time"] = df["time"] / 1_000
        
        for k, v in parse_metadata_content(metadata, None):
            df.attrs[k] = v

    df.attrs[ATTR_TIME_UNITS] = "microseconds"
    df.attrs[ATTR_SIGNAL_UNITS] = "volts"
    df.attrs[ATTR_REPEATS] = None
    return df


def read_with_repeats(p: pathlib.Path | str) -> DataFrame:
    """Read Edinburgh Instruments ascii file (with repeats).
    """

    if isinstance(p, str):
        p = pathlib.Path(p)
        
    with p.open("r", encoding="ascii") as fi:
        fi.readline() # In files with repeats, the first two lines are like a header
        fi.readline() # In files with repeats, the first two lines are like a header

        metadata = read_metadata(fi)

        # As the number of repeats is not stored in the metadata,
        # we infer this value by counting the number of values in 
        # one the metadata keys.

        estimated_length = len(metadata["Averages"].split(","))

        # TODO: is the time always in milliseconds? is the signal always in Volts?
        df = pd.read_csv( # type: ignore
            fi, 
            sep=",",
            header=0,
            names=("time", ) + tuple("signal%d" % n for n in range(estimated_length))
        )
        df["time"] = df["time"] / 1_000
        
        for k, v in parse_metadata_content(metadata, estimated_length):
            assert len(v) == estimated_length
            df.attrs[k] = v

    df.attrs[ATTR_TIME_UNITS] = "microseconds"
    df.attrs[ATTR_SIGNAL_UNITS] = "volts"
    df.attrs[ATTR_REPEATS] = estimated_length
    return df


def read(p: pathlib.Path | str) -> DataFrame:
    """Read Edinburgh Instruments ascii file (with or without repeats).
    """

    if isinstance(p, str):
        p = pathlib.Path(p)

    # is there a better way?
    with p.open("r", encoding="ascii") as fi:
        # In a file with repeats, the first line contains the filename.
        # In a file without repeats, the file line contains the 
        # first metadata key value pair, comma separated.
        # This will fail is the filename contains a comma (which is rare)

        if "," in fi.readline():
            return read_without_repeats(p)
        else:
            return read_with_repeats(p)


def yield_individual_repeats(raw_df: DataFrame) -> Generator[tuple[int | None, DataFrame], None, None]:
    """Yield number of repetition and dataframe.
    """

    time: Array = raw_df["time"].to_numpy()

    if ATTR_REPEATS in raw_df.attrs and raw_df.attrs[ATTR_REPEATS] is not None:    
        for ndx in range(raw_df.attrs[ATTR_REPEATS]):
            signal: Array = raw_df["signal%d" % ndx].to_numpy()
            tmpdf = pd.DataFrame(dict(time=time, signal=signal))
            for k, v in raw_df.attrs.items():
                if isinstance(v, (tuple, ndarray)):
                    tmpdf.attrs[k] = v[ndx]
                else:
                    tmpdf.attrs[k] = v
            yield ndx, tmpdf
    else:
        yield None, raw_df


##############
# GUI Helpers
##############

def open_explorer(path: pathlib.Path | str):
    """Open the OS file explorer at the given path.
    """

    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


#####################
# Plotting functions
#####################

# Internal variable to footnote timestamp.
_footnote_timestamp: str | None = None

def footnote(fig: Figure, *, left_footer: str = "", right_footer: str = ""):
    """Add footnote to page.
    """

    if left_footer:
        fig.text(0.02, 0.01, left_footer, ha='left', fontsize=6, wrap=True) # type: ignore
    if right_footer:
        fig.text(0.98, 0.01, right_footer, ha='right', fontsize=6, wrap=True) # type: ignore
    

def default_footnote(fig: Figure | None):
    """Add default footnote to page, which includes the analysis
    datetime and the script version.

    To initialize the analysis datetime to current time, 
    call this function with None value.
    """
    global _footnote_timestamp
    if fig is None:
        _footnote_timestamp = datetime.datetime.now().isoformat(timespec="seconds")
    else:
        footnote(
            fig, 
            left_footer=f"Analysis datetime: {_footnote_timestamp}",
            right_footer=f"Photoacoustic analysis version: {__version__}",
        )


def plot_signal_and_peaks(ax: Axes, raw_df: DataFrame, trace_analysis: TraceAnalysis, signal_smooth: Array | None):
    """Plot the signal and peaks (if given).

    Parameters
    ----------
    ax
        Matplotlib axes to draw.
    signal_df
        Dataframe with the signal.
    peak_df
        Dataframe with found peaks.
    """

    ax.plot(raw_df["time"], raw_df["signal"], c="tab:gray")

    if signal_smooth is not None:
        ax.plot(raw_df["time"], signal_smooth, c="tab:blue")
    
    mx = np.max(np.abs((ax.get_ylim())))
    ax.set_ylim(-mx, mx)

    bg = raw_df["signal"][:500].mean() 
    std = raw_df["signal"][:500].std()

    ax.axhline(y=bg, ls=':', c="black")
    ax.axhline(y=bg - std, ls=':', c="black")
    ax.axhline(y=bg + std, ls=':', c="black")

    ax.set_xlabel(r"time / $\mu s$")
    ax.set_ylabel("signal / V")

    for n in (1, 2):
        x = trace_analysis["time_peak1"].nominal_value if n == 1 else trace_analysis["time_peak2"].nominal_value
        y = trace_analysis["signal_peak1"].nominal_value if n == 1 else trace_analysis["signal_peak2"].nominal_value
        if np.isnan(x) or np.isnan(y):
            continue
        ax.axvline(x=x, ls='--', c="tab:green")
        ax.axhline(y=y, ls='--', c="tab:green")
        ax.plot([x], [y], "x", c="tab:red")


def build_time_trace_figure(
    raw_df: DataFrame, 
    peak_record: TraceAnalysis, 
    signal_smooth: Array,
    ) -> Figure:
    """Plot a figure 

    Parameters
    ----------
    sample
        name of the sample.
    signal_df
        Dataframe with the signal.
    peak_df
        Dataframe with the found peaks.
    title, optional
        
    """

    fig = plt.figure()
    fig.set_figwidth(297/40)
    fig.set_figheight(210/40)

    gs = GridSpec(3, 3, figure=fig)#, bottom=.05)
    ax_plot = fig.add_subplot(gs[:2, :])
    ax_meta = fig.add_subplot(gs[2, :2])
    ax_peak = fig.add_subplot(gs[2, -1])

    ax_plot.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.3f"))

    ax_inset: Axes = inset_axes(ax_plot, width="30%", height="20%")
    ax_inset.plot(raw_df["time"], raw_df["signal"])
    ax_inset.get_xaxis().set_ticks([])
    ax_inset.get_yaxis().set_ticks([])
    
    plot_signal_and_peaks(ax_plot, raw_df, peak_record, signal_smooth)
    ax_meta.axis(False)
    ax_peak.axis(False)
    
    cellText = [
            ("Description", raw_df.attrs["Desc"]),
            ("Wavelength", f"{raw_df.attrs['Wavelength']} nm"),
            ("Laser energy", r"$({:.2uL})~\mu J$".format(raw_df.attrs['Laser energy before'])),
            ("Exc. Wavelength", f'{raw_df.attrs.get("exc_wavelength", "N/A")} nm'),
        ]

    table = ax_meta.table(
        cellText=cellText,
        colLabels=("Parameter", "Value"),
        loc='center'
    )
    table.auto_set_font_size(False)
    table.set_fontsize(5)

    ax_peak.set_title("Peaks")
    table = ax_peak.table(
        cellText=[
            (f"{peak_record['time_peak1'].nominal_value:.3f}", f"{peak_record['signal_peak1'].nominal_value:.3f}"),
            (f"{peak_record['time_peak2'].nominal_value:.3f}", f"{peak_record['signal_peak2'].nominal_value:.3f}"),
            (f"{peak_record['time_delta'].nominal_value:.3f}", f"{peak_record['signal_delta'].nominal_value:.3f}"),
        ],
        colLabels=(
            #"Peak #", 
            r"Time / $\mu s$", 
            "Signal / V"),
        rowLabels=(
            " #1 ",
            " #2 ",
            r" $\Delta$ "
        ),
        loc='center',
    )
    table.auto_set_font_size(False)
    table.set_fontsize(5)
    # table.scale(1, 4)
        
    t0 = peak_record["time_peak1"]
    t1 = peak_record["time_peak2"]
    if np.isfinite(t0.nominal_value) and np.isfinite(t1.nominal_value):
        lb = t0.nominal_value - 3 * (t1.nominal_value - t0.nominal_value)
        ub = t0.nominal_value + 6 * (t1.nominal_value - t0.nominal_value)
        ax_plot.set_xlim(lb, ub)
        ax_inset.axvline(x=lb, ls="-", c="black")
        ax_inset.axvline(x=ub, ls="-", c="black")

    fig.tight_layout()

    return fig


def build_powerscan_overview_figure(signals: list[tuple[Array, Array]], energy: Array) -> Figure:

    fig, ax = plt.subplots(1, 1)

    fig.set_figwidth(297/40)
    fig.set_figheight(210/40)

    ax.set_xlabel(r"$\Delta$time / $\mu s$")
    ax.set_ylabel("signal / V")
    ax.set_xlim(-2, 4)

    try:
        norm = colors.Normalize(vmin=energy.min(), vmax=energy.max())
    except Exception:
        print(energy)
        norm = colors.Normalize(vmin=0, vmax=1)

    for power, (time, signal) in zip(energy, signals):
        ax.plot(time, signal, c=plt.cm.jet(norm(power)))

    plt.colorbar(
        plt.cm.ScalarMappable(norm=norm, cmap=plt.cm.jet), 
        orientation='vertical', 
        ax=ax, 
        label=r"Laser power / $\mu J$"
    )
    default_footnote(fig)
    plt.tight_layout()

    return fig


def build_powerscan_figure(
        energy_delta_signal: Iterable[tuple[Iterable[Variable], Iterable[Variable]]],
        slope_intercepts: tuple[Iterable[Variable], Iterable[Variable]],
        slopes0: Iterable[Variable],
        labels: Iterable[str],
        ) -> Figure:
    """_summary_

    Parameters
    ----------
    wavelength
        _description_
    pdf
        _description_

    Yields
    ------
        _description_
    """

    fig, (ax_plot, ax_meta) = plt.subplots(
        2, 1, 
        gridspec_kw=dict(height_ratios=(.7, .3))
    )
    fig.set_figwidth(297/40)
    fig.set_figheight(210/40)

    ax_plot.set_xlabel(r"Laser power / $\mu J$")
    ax_plot.set_ylabel(r"$\Delta$signal / V")
    ax_plot.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.3f"))

    cellText = []
    rowLabels = []
    rowColours = []

    for (x, y), slope, intercept, slope0, label in zip(energy_delta_signal, *slope_intercepts, slopes0, labels):
        x, x_unc = split_unc_tuple(*x)
        y, y_unc = split_unc_tuple(*y)

        x_fit = np.linspace(0, np.max(x) * 1.1, 10)
        y_fit = slope.nominal_value * x_fit + intercept.nominal_value
        line, = ax_plot.plot(x_fit, y_fit)

        x_fit = np.linspace(0, np.max(x) * 1.1, 10)
        y_fit = slope0.nominal_value * x_fit
        line, = ax_plot.plot(x_fit, y_fit, ls=":", color=line.get_color())

        ax_plot.errorbar(
            x, 
            y, 
            xerr=x_unc, 
            yerr=y_unc, 
            linestyle='None', 
            marker='.',
            color=line.get_color(),
        )

        cellText.append(
            (label, f"${slope:.2uL}$", f"${intercept:.2uL}$", f"${slope0:.2uL}$"),
        )

        rowColours.append(line.get_color())
        rowLabels.append("   ")

    table = ax_meta.table(
        cellText=cellText,
        colLabels=("Subfolder", r"Slope / $\left(V / \mu J \right)$", "Intercept / $V$", r"Slope0 / $\left(V / \mu J \right)$"),
        loc='center',
        rowColours=rowColours,
        rowLabels=rowLabels,
    )

    table.auto_set_font_size(False)
    table.set_fontsize(5)

    ax_meta.axis(False)

    default_footnote(fig)
    fig.tight_layout()

    return fig


#####################
# Analysis functions
#####################

def find_first_two_peaks(time: Array, signal: Array, signal_smooth: Array | None) -> list[tuple[Variable, Variable]]:
    """Find upto first two peaks.

    Iterable of Time, Signal
    """
    bg = signal[:500].mean() 
    std = signal[:500].std()
    
    # MHz
    acq_frequency = 1 / np.diff(time)[0]
    # us
    time_distance = 1/2
    time_width = 1/4

    if signal_smooth is None:
        signal_smooth = savgol_filter(signal, 51, 3)

    out = []

    # Find positive peaks

    ndxs, _props = find_peaks(
        -signal_smooth, 
        height=-bg + 2 * std,
        prominence=2*std,
        distance=time_distance * acq_frequency,
        width=time_width * acq_frequency,
    )

    out.append(
        pd.DataFrame(
            dict(
                time=time[ndxs], 
                time_unc=np.zeros_like(time[ndxs]),
                signal=signal_smooth[ndxs],
                signal_unc=np.zeros_like(time[ndxs]),
            )
        )
    )

    # Find negative peaks peaks

    ndxs, _props = find_peaks(
        signal_smooth, 
        height=bg + 2 * std,
        prominence=2*std,
        distance=time_distance * acq_frequency,
        width=time_width * acq_frequency,
    )

    out.append(
        pd.DataFrame(
            dict(
                time=time[ndxs], 
                time_unc=np.zeros_like(time[ndxs]),
                signal=signal_smooth[ndxs],
                signal_unc=np.zeros_like(time[ndxs]),
            )
        )
    )
    
    out = pd.concat(out).sort_values(by="time").reset_index(drop=True)

    if len(out) == 0:
        return []
    
    # Keep only the first two peaks: positive, followed by negative.

    delta = np.abs(out["signal"].values - bg)
    sel = delta > np.max(delta) / 20 + std
    out = out[sel].reset_index(drop=True)

    sign = np.sign(out["signal"].values - bg)
    sel = np.zeros(len(out), dtype=bool)
    best = np.where((sign == 1) & (np.roll(sign, -1) == -1))[0]
    
    if len(best) == 0:
        return []
    
    best = best[0]
    sel[best] = True
    sel[best + 1] = True

    return [ ( ufloat(out.iloc[ndx]["time"], out.iloc[ndx]["time_unc"]), 
               ufloat(out.iloc[ndx]["signal"], out.iloc[ndx]["signal_unc"]) ) 
              for ndx in (best, best+1)]


def _fix_fit_unc(unc: Array):
    if not np.any(unc == 0):
        return unc
    
    if np.all(unc == 0):
        return None
    else:
        unc = np.copy(unc)
        unc[unc == 0] = unc[unc > 0].min() / 10
        return unc


def fit_linear(x: Iterable[float], y: Iterable[float], x_unc: Iterable[float], y_unc: Iterable[float], intercept0: bool=False) -> tuple[Variable, Variable]:
    """Fit linear and return the slope and intercept (as value with uncertainty).
    """

    x = np.asarray(x)
    y = np.asarray(y)
    x_unc = np.asarray(x_unc)
    y_unc = np.asarray(y_unc)

    data = odr.RealData(
        x, y,
        sx=_fix_fit_unc(x_unc), sy=_fix_fit_unc(y_unc),
    )

    if intercept0:
        beta0 = [np.median(y_unc[x_unc>0]/x_unc[x_unc>0]), 0]
        result = odr.ODR(data, odr.unilinear, beta0=beta0, ifixb=[1, 0]).run()
    else:
        result = odr.ODR(data, odr.unilinear).run()

    return to_unc_tuple(result.beta, result.sd_beta)


def analyze_time_trace(time: Array, signal: Array, options: Options) -> tuple[TraceAnalysis, Array]:
    """Find the first two peaks to obtain the time and signal delta.
    """

    signal_smooth: Array = savgol_filter(signal, options["savgol_window_length"], options["savgol_polyorder"])

    peaks = find_first_two_peaks(
        time, 
        signal, 
        signal_smooth
    )
    
    # TODO: Just in case
    peaks.append((UFLOAT_NAN, UFLOAT_NAN))
    peaks.append((UFLOAT_NAN, UFLOAT_NAN))

    time_delta  = peaks[0][0] - peaks[1][0]
    signal_delta = peaks[0][1] - peaks[1][1]

    return {
        "energy": UFLOAT_NAN,

        "time_peak1": peaks[0][0],
        "signal_peak1": peaks[0][1],

        "time_peak2": peaks[1][0],
        "signal_peak2": peaks[1][1],

        "time_delta": time_delta,
        "signal_delta": signal_delta,
    }, signal_smooth
    

def analyze_file(p: pathlib.Path, pdf: PdfPages | None, xlsx: pd.ExcelWriter | None, options: Options) -> tuple[FileAnalysis, list[tuple[Array, Array]], list[float]]:
    """Analyze a file (with or without repeats) to obtain the time and signal delta.
    """

    experiment_folder = p.parent.parent

    try:
        alldf = read(p)
    except Exception as ex:
        options["on_error"](f"Could not load file {str(p.relative_to(experiment_folder))}: {str(ex)}")
        return {}, [], [] # type: ignore
    
    if not len(alldf):
        return {}, [], [] # type: ignore
    
    options["on_progress"](str(p.relative_to(experiment_folder)))

    signals: list[tuple[Array, Array]] = []

    records: list[TraceAnalysis] = []

    for ndx, df in yield_individual_repeats(alldf):
        suffix = "" if ndx is None else f"\n(rep {ndx+1}/{df.attrs[ATTR_REPEATS]})"
        
        try:
            trace_analysis, signal_smooth = analyze_time_trace(df["time"].to_numpy(), df["signal"].to_numpy(), options)
        except Exception as ex: 
            options["on_error"](f"Could not analyze time trace {str(p.relative_to(experiment_folder))} {suffix}: {str(ex)}")
            continue

        if not np.isnan(trace_analysis["time_peak1"].nominal_value):
            signals.append((df["time"].to_numpy() - trace_analysis["time_peak1"].nominal_value, signal_smooth))

        trace_analysis["energy"] = df.attrs["Laser energy before"]
        records.append(trace_analysis)

        if pdf is not None:
            if ndx is None or options["plot_time_trace_rep"]:
                fig = build_time_trace_figure(
                        df, trace_analysis, signal_smooth,
                    )
                plt.suptitle(str(p.relative_to(experiment_folder)) + suffix)
                default_footnote(fig)
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)


    trace_analysis_df = pd.DataFrame.from_records(records)

    if ndx is not None and xlsx:
        # Only store the file sheet, if repeats were used.
        try:
            prefix = "(%s)" % p.parent.stem.split("_")[0]
        except Exception:
            prefix = "(?)"
        
        unzip_unc_column(trace_analysis_df.copy(), *get_unc_keys(TraceAnalysis), drop_unc=True).to_excel(
            xlsx, 
            sheet_name= prefix + " " + p.stem, 
            startrow=0,
            index=False,
            header=True
        )

    #####################
    # Results of repeat

    try:
        # TODO: check what Edinburg is doing for compatibility std or sem
        # TODO: make funciton filter all simulteanously.
        valid = [np.isfinite(delta.nominal_value) and np.isfinite(delta.std_dev)
                 for delta in trace_analysis_df["time_delta"].values]

        energy = ufloat_nanmean(*trace_analysis_df[valid]["energy"].to_list())
        time_delta = ufloat_nanmean(*trace_analysis_df[valid]["time_delta"].to_list())
        signal_delta = ufloat_nanmean(*trace_analysis_df[valid]["signal_delta"].to_list())
    except Exception as ex:
        print(ex)
        energy  = time_delta = signal_delta = UFLOAT_NAN

    return {
        "path": str(p.relative_to(experiment_folder)),
        "description": df.attrs["Desc"],
        "comment": df.attrs["Comment"],
        "wavelength": df.attrs["Wavelength"],
        "bandwith": df.attrs["Bandwidth"],
        "averages": df.attrs["Averages"],
        "repeats": df.attrs[ATTR_REPEATS],

        "energy": energy,

        "time_delta": time_delta,
        "signal_delta": signal_delta,
    }, signals, [x.nominal_value for x in trace_analysis_df["energy"]]



def analyze_powerscan_folder(folder: pathlib.Path, pdf: PdfPages | None, xlsx: pd.ExcelWriter | None, options: Options) -> tuple[PowerscanAnalysis, tuple[list[Variable], list[Variable]]]:
    """Analyze a powerscan folder to the slope and intercept
    of the delta signal vs energy.
    """

    parts = folder.stem.split("_", 2)
    if len(parts) == 3:
        sam_ref, wl, desc = parts
        try:
            wl = float(wl)
        except Exception:
            print(f"{folder}: wavelength must be a number")
            wl = np.nan
    else:
        print(f"{folder}: invalid")
        sam_ref, wl, desc = "N/A", np.nan, folder.stem

    records: list[FileAnalysis] = []

    signals: list[tuple[Array, Array]] = []
    powers: list[float] = []
    
    median_signals: list[tuple[Array, Array]] = []
    median_powers: list[float] = []

    for file in folder.glob("*.txt"):
        if file.stem.startswith("_"):
            continue

        file_analysis, _signals, _powers= analyze_file(file, pdf, xlsx, options)

        if len(_signals) == 0:
            continue
        elif len(_signals) > 1:
            ndx = _argmedian_at_t0(_signals)
        else:
            ndx = 0

        median_signals.append(_signals[ndx])
        median_powers.append(_powers[ndx])

        signals.extend(_signals)
        powers.extend(_powers) 
        records.append(file_analysis)

    file_df = pd.DataFrame.from_records(records)

    if pdf is not None:
        if len(signals) > len(median_signals):
            fig = build_powerscan_overview_figure(signals, np.asarray(powers))
            fig.suptitle(folder.stem)
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
        
        fig = build_powerscan_overview_figure(median_signals, np.asarray(median_powers))
        fig.suptitle(folder.stem)
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    if xlsx is not None:
        unzip_unc_column(file_df.copy(), *get_unc_keys(FileAnalysis)).to_excel(
            xlsx, 
            sheet_name=folder.stem, 
            index=False
        )

    energy: list[Variable] = file_df["energy"].to_list()
    signal_delta: list[Variable] = file_df["signal_delta"].to_list()

    try:
        x, x_unc = split_unc_tuple(*energy, container=lambda el: np.fromiter(el, dtype=float))
        y, y_unc = split_unc_tuple(*signal_delta, container=lambda el: np.fromiter(el, dtype=float))

        valid = np.logical_not(np.logical_or(np.isnan(x), np.isnan(y)))
        if np.sum(valid) >= 2:
            slope, intercept = fit_linear(x[valid], y[valid], x_unc[valid], y_unc[valid])
            slope0, _intercept0 = fit_linear(x[valid], y[valid], x_unc[valid], y_unc[valid], intercept0=True)
        else:
            options["on_error"]("Could not fit for {folder.stem}: not enough valid points")
            slope = intercept = UFLOAT_NAN
            slope0 = _intercept0 = UFLOAT_NAN

    except Exception as ex:
        options["on_error"](f"Could not fit for {folder.stem}: {str(ex)}")
        slope = intercept = UFLOAT_NAN
        slope0 = _intercept0 = UFLOAT_NAN


    return {
        "folder": folder.stem,
        "sam_ref": sam_ref,
        "exc_wavelength": wl,
        "description": desc,
        "slope": slope,
        "intercept": intercept,
        "slope0": slope0,
    }, (energy, signal_delta)


def analyze_experiment_folder(folder: pathlib.Path, pdf: PdfPages | None, xlsx: pd.ExcelWriter | None, options: Options) -> DataFrame:
    """Analyze experiment folder to obtain the alpha value for the sample.
    """

    abs_ref = None
    abs_sam = None
    try:
        for line in (folder / "abs.txt").read_text().splitlines():
            k, v = line.split("=")
            k = k.strip().lower()
            if k == "ref":
                abs_ref = float(v.strip())
            elif k == "sam":
                abs_sam = float(v.strip())
        if abs_sam is None:
            options["on_error"]("Sample absorption not found in `abs.txt`")
        if abs_ref is None :
            options["on_error"]("Reference absorption not found in `abs.txt`")
    except Exception as ex:
        options["on_error"](f"Absorption values could not be loaded: {str(ex)}")

    records: list[PowerscanAnalysis] = []

    xys: list[tuple[list[Variable], list[Variable]]] = []

    for subfolder in folder.iterdir():
        if not subfolder.is_dir():
            print(f"{subfolder}: skipping, not a folder.")
            continue
        if subfolder.stem.startswith("_"):
            print(f"{subfolder}: skipping, user skip prefix.")
            continue

        powerscan_analysis, xy = analyze_powerscan_folder(subfolder, pdf, xlsx, options)

        records.append(powerscan_analysis)
        xys.append(xy)

    fit_df = pd.DataFrame.from_records(records)

    for exc_wavelength, gdf in fit_df.groupby("exc_wavelength"):
        try:
            row_ref  = gdf.query("sam_ref.str.startswith('ref')").sort_values("sam_ref").iloc[0]
        except Exception:
            row_ref = None

        if row_ref is not None:
            tmp = []
            tmp0 = []
            folders = []
            for _, row in gdf.iterrows():
                folders.append(row["folder"])
                if row["folder"] == row_ref["folder"]:
                    tmp.append("")
                    tmp0.append("")
                else:
                    tmp.append(ztest(row["slope"], row_ref["slope"]))
                    tmp0.append(ztest(row["slope0"], row_ref["slope0"]))
            gdf["p-value " + row_ref["sam_ref"] + " slope"] = tmp
            gdf["p-value " + row_ref["sam_ref"] + " slope0"] = tmp0

            del tmp, tmp0, folders

        if pdf is not None:
            fig = build_powerscan_figure(
                xys, 
                (gdf["slope"].to_list(), gdf["intercept"].to_list()),
                gdf["slope0"].to_list(),
                gdf["folder"].to_list()
            )
            fig.suptitle(f"Excitation Wavelength {exc_wavelength} nm")
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        if xlsx is not None:
            unzip_unc_column(gdf, "slope", "intercept", "slope0").to_excel(
                xlsx, sheet_name=f"FIT {exc_wavelength}", index=False
            )

    if abs_ref is None or abs_sam is None:
        return pd.DataFrame()
    
    # (m_sam / m_ref) = alpha * (1- 10^-(A_sam)) / (1- 10^-(A_ref))
    alpha_records: list[ExperimentAnalysis] = []

    factor = (1 - 10**(-abs_ref)) / (1 - 10**(-abs_sam))

    for exc_wavelength, gdf in fit_df.groupby("exc_wavelength"):

        gdf_ref = gdf.query("sam_ref.str.startswith('ref')")
        for _, row_sam in gdf.query("sam_ref.str.startswith('sam')").iterrows():
            for _, row_ref in gdf_ref.iterrows():
                alpha_records.append(
                    {
                        "abs sam": abs_sam,
                        "abs_ref": abs_ref,
                        "ref": row_ref["folder"],
                        "sam": row_sam["folder"],
                        "exc_wavelength": exc_wavelength,
                        "alpha": row_sam["slope"] / row_ref["slope"] * factor,
                        "alpha0": row_sam["slope0"] / row_ref["slope0"] * factor,
                    }
                )


        refs = gdf_ref["slope"].to_list()
        refs0 = gdf_ref["slope0"].to_list()
        if len(refs):
            alpha_records.append(
                {
                    "abs sam": abs_sam,
                    "abs_ref": abs_ref,
                    "ref": "avg",
                    "sam": row_sam["folder"],
                    "exc_wavelength": exc_wavelength,
                    "alpha": row_sam["slope"] / ufloat_nanmean(*refs) * factor,
                    "alpha0": row_sam["slope0"] / ufloat_nanmean(*refs0) * factor,
                }
            )

    return pd.DataFrame.from_records(alpha_records)


def analyze(root: pathlib.Path, options: Options | None=None):
    """Analyze an experiment folder:

    1. analyze all sample and reference folders.
    2. save a pdf and xlsx file with the results.
    """

    if options is None:
        options = DEFAULT_OPTIONS
    else:
        options = {**DEFAULT_OPTIONS, **options}
    assert options is not None

    default_footnote(None)

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="This figure includes Axes that are not compatible with tight_layout")
        with PdfPages(root / 'summary.pdf') as pdf:

            d = pdf.infodict()
            d['Title'] = 'Photoacoustic analysis'
            d['Author'] = 'Hernán Grecco'
            d['CreationDate'] = datetime.datetime.today()
            d['ModDate'] = datetime.datetime.today()
            
            with pd.ExcelWriter(root / 'summary.xlsx') as xlsx:
                df = analyze_experiment_folder(root, pdf, xlsx, options) 
                if len(df):
                    unzip_unc_column(df, "alpha", "alpha0").to_excel(xlsx, sheet_name="__ALPHA__", index=False)

            reorganize_sheets(root / 'summary.xlsx')


if __name__ == "__main__":
    # from tkinter import Tk
    # from tkinter import filedialog
    # path = pathlib.Path(filedialog.askdirectory(initialdir="."))
    # root = Tk()
    ROOT = pathlib.Path("/Users/grecco/Documents/projects/strassert/optoacustic/data") 

    # path = ROOT / "2024-07-16"
    # analyze(path)
    # path = ROOT / "Short scale"
    # analyze(path)
    # path = ROOT / "2024-08-08"
    # analyze(path)
    # path = ROOT / "2024-08-01" / "Ar" / "Averages"
    # analyze(path)
    # path = ROOT / "2024-08-01" / "Ar" / "10 Measurements"
    # analyze(path)
    # path = ROOT / "2024-08-01" / "Air" / "Averages"
    # analyze(path)
    # path = ROOT / "2024-08-01" / "Air" / "10 Measurements"
    # analyze(path)
    path = ROOT / "2024-08-09"
    analyze(path)
    # open_explorer(ROOT)
    # root.mainloop()